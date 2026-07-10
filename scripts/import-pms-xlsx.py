"""PMS-ENGINE.xlsx + PMS-DECK.xlsx → data/pms-unified.json"""
import json
import re
import uuid
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DESKTOP = Path(r"c:\Users\SAMSUNG\OneDrive\바탕 화면")
ENGINE_PATH = DESKTOP / "PMS-ENGINE.xlsx"
DECK_PATH = DESKTOP / "PMS-DECK.xlsx"
OUT_PATH = ROOT / "data" / "pms-unified.json"

HEADER_MAP = {
    "GROUP": "group",
    "JOB CODE": "job_code",
    "SORT": "sort",
    "MAINTENANCE ITEM (SORT-1)": "item_sort1",
    "MAINTENANCE ITEM (SORT-2)": "item_sort2",
    "JOB DETAIL": "job_detail",
    "PERIOD": "period",
    "UNIT": "unit",
    "P.I.C": "pic",
    "LASTDONE": "last_done",
    "HISTORY": "history",
    "REMARK": "remark",
}


def norm_header(h):
    if h is None:
        return ""
    s = str(h).replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if s.startswith("NEXT DATE"):
        return "next_date"
    if s.startswith("LAST DONE") or s == "LAST DONE":
        return "last_done"
    return HEADER_MAP.get(s.upper().replace("  ", " "), s)


def parse_sheet(wb, sheet_name, department):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm_header(c) for c in rows[0]]
    # DECK col6 blank header → job_detail
    for i, h in enumerate(headers):
        if h in ("", "   ") or (isinstance(h, str) and not h.strip()):
            headers[i] = "job_detail"
    jobs = []
    for row in rows[1:]:
        if not row or not row[1]:
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h or i >= len(row):
                continue
            val = row[i]
            if val is None:
                continue
            if h == "period":
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    continue
            elif h in ("next_date", "last_done") and hasattr(val, "isoformat"):
                val = val.strftime("%Y-%m-%d")
            else:
                val = str(val).strip()
            rec[h] = val
        if not rec.get("job_code"):
            continue
        rec["id"] = str(uuid.uuid4())
        rec["department"] = department
        rec["is_overdue"] = _calc_overdue(rec.get("next_date"))
        jobs.append(rec)
    return jobs


def _calc_overdue(next_date):
    if not next_date:
        return False
    try:
        nd = datetime.strptime(str(next_date)[:10], "%Y-%m-%d")
        return nd.date() < datetime.now().date()
    except ValueError:
        return False


def build_component_tree(jobs):
    components = {}
    order = 0

    def ensure(path, node_type, label, parent_id=None):
        nonlocal order
        key = "|".join(path)
        if key in components:
            return components[key]["id"]
        cid = str(uuid.uuid4())
        order += 1
        components[key] = {
            "id": cid,
            "parent_id": parent_id,
            "path": path,
            "label": label,
            "node_type": node_type,
            "sort_order": order,
        }
        return cid

    for job in jobs:
        dept = job["department"]
        parts = [dept, job.get("group", ""), job.get("sort", ""), job.get("item_sort1", ""), job.get("item_sort2", "")]
        parts = [p.strip() for p in parts if p and str(p).strip()]
        parent = None
        path_acc = []
        types = ["DEPARTMENT", "GROUP", "SORT", "ITEM_L1", "ITEM_L2"]
        for i, p in enumerate(parts):
            path_acc.append(p)
            parent = ensure(path_acc, types[min(i, len(types) - 1)], p, parent)
        job["ship_component_id"] = parent

    return list(components.values())


def default_spares():
    return [
        {"id": str(uuid.uuid4()), "part_no": "ME-EX-001", "name": "Exhaust Valve Spindle", "qty_on_hand": 2, "min_qty": 2, "unit": "EA"},
        {"id": str(uuid.uuid4()), "part_no": "ME-FI-012", "name": "Fuel Injector Nozzle Tips", "qty_on_hand": 4, "min_qty": 6, "unit": "EA"},
        {"id": str(uuid.uuid4()), "part_no": "AE-PR-04", "name": "Piston Ring Set", "qty_on_hand": 3, "min_qty": 2, "unit": "EA"},
    ]


def main():
    engine_jobs = []
    deck_jobs = []
    if ENGINE_PATH.exists():
        wb = openpyxl.load_workbook(ENGINE_PATH, data_only=True)
        engine_jobs = parse_sheet(wb, "ENGINE", "ENGINE")
    if DECK_PATH.exists():
        wb = openpyxl.load_workbook(DECK_PATH, data_only=True)
        deck_jobs = parse_sheet(wb, "DECK", "DECK")

    all_jobs = engine_jobs + deck_jobs
    components = build_component_tree(all_jobs)

    payload = {
        "meta": {
            "vessel_id": "TEST_V01",
            "imported_at": datetime.now().isoformat(),
            "engine_count": len(engine_jobs),
            "deck_count": len(deck_jobs),
        },
        "ship_components": components,
        "maintenance_jobs": all_jobs,
        "spare_parts": default_spares(),
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(all_jobs)} jobs, {len(components)} components → {OUT_PATH}")


if __name__ == "__main__":
    main()
